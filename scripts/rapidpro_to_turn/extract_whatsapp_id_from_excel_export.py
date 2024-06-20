import argparse

from openpyxl import load_workbook


def get_arguments():
    parser = argparse.ArgumentParser(
        description="Extracts the whatsapp IDs from a rapidpro contacts export"
    )
    parser.add_argument("filename", help="Excel file to extract from")
    return parser.parse_args()


def main():
    args = get_arguments()
    wb = load_workbook(args.filename, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == "URN:Whatsapp":
                urn_col = cell.column
                break
        break
    for row in ws.iter_rows(min_row=2):
        whatsapp_id = row[urn_col - 1].value
        if whatsapp_id and whatsapp_id.isnumeric():
            print(whatsapp_id)


if __name__ == "__main__":
    main()
