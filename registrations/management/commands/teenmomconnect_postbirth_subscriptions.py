from collections import namedtuple
from functools import lru_cache

import phonenumbers
from django.core.management.base import BaseCommand, CommandError
from iso639 import languages
from openpyxl import load_workbook

from ndoh_hub.utils import LANGUAGES, is_client, sbm_client
from registrations.models import SubscriptionRequest

Contact = namedtuple("Contact", ["msisdn", "language"])


class Command(BaseCommand):
    help = (
        "Creates post birth subscriptions for the mothers specified in the xlsx file. "
        "Doesn't create subscriptions for mothers that already have an active "
        "subscription."
    )

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="The path to the xlsx file")

    @staticmethod
    def extract_contacts_from_workbook(file_path):
        """
        Extracts the list of Contacts from the workbook at `file_path`, raising a
        CommandError if there are any issues with the workbook's format
        """
        workbook = load_workbook(file_path, read_only=True)

        try:
            [worksheet] = workbook.worksheets
        except ValueError:
            raise CommandError("Document can only have a single sheet")

        try:
            [phone_header] = filter(lambda r: r.value == "Phone", worksheet[1])
        except (ValueError, IndexError):
            raise CommandError("Sheet must have a single 'Phone' column header")
        try:
            [language_header] = filter(lambda r: r.value == "Language", worksheet[1])
        except (ValueError, IndexError):
            raise CommandError("Sheet must have a single 'Language' column header")

        for row in worksheet.iter_rows(min_row=2):
            yield Contact(
                row[phone_header.column - 1].value,
                row[language_header.column - 1].value,
            )

    def validate_msisdn(self, msisdn):
        """
        Returns an E164 internationally formatted msisdn string if valid, else logs an
        error and returns `None`
        """
        try:
            p = phonenumbers.parse(msisdn, "ZA")
            assert phonenumbers.is_possible_number(p)
            assert phonenumbers.is_valid_number(p)
            return phonenumbers.format_number(p, phonenumbers.PhoneNumberFormat.E164)
        except (phonenumbers.phonenumberutil.NumberParseException, AssertionError):
            self.stdout.write(
                self.style.NOTICE("Invalid phone number {}. Skipping...".format(msisdn))
            )
            return None

    def validate_language(self, language):
        """
        Returns the language in the Seed language format if valid, else logs an error
        and returns `None`
        """
        try:
            languages.get(part3=language)
            seed_lang = "{}_ZA".format(language)
            assert seed_lang in LANGUAGES
            return seed_lang
        except (KeyError, AssertionError):
            self.stdout.write(
                self.style.NOTICE(
                    "Invalid language code {}. Skipping...".format(language)
                )
            )
            return None

    @staticmethod
    def create_or_update_identity(msisdn, language):
        """
        Fetches the identity with the given msisdn, ensures that the language code is
        correct, or if identity doesn't exist, creates and returns it.
        """
        try:
            identity = is_client.get_identity_by_address("msisdn", msisdn)["results"]
            identity = next(identity)
        except StopIteration:
            identity = is_client.create_identity(
                {
                    "details": {
                        "lang_code": language,
                        "default_addr_type": "msisdn",
                        "addresses": {
                            "msisdn": {msisdn: {"default": True, "optedout": False}}
                        },
                    }
                }
            )
        if identity["details"].get("lang_code") != language:
            identity["details"]["lang_code"] = language
            identity = is_client.update_identity(
                identity["id"], {"details": identity["details"]}
            )
        return identity

    @staticmethod
    @lru_cache()
    def _get_messageset(short_name):
        [messageset] = sbm_client.get_messagesets(params={"short_name": short_name})[
            "results"
        ]
        return messageset

    def create_subscription_postbirth(self, identity):
        """
        Creates a subscription to the postbirth messageset for `identity` if the
        identity doesn't have any active subscriptions. Logs an error if the identity
        already has an active subscription
        """
        try:
            subscriptions = sbm_client.get_subscriptions(
                params={"identity": identity["id"], "active": True}
            )["results"]
            next(subscriptions)
        except StopIteration:
            messageset = self._get_messageset("momconnect_postbirth.hw_full.1")
            sr = SubscriptionRequest.objects.create(
                identity=identity["id"],
                messageset=messageset["id"],
                lang=identity["details"]["lang_code"],
                schedule=messageset["default_schedule"],
            )
            self.stdout.write(
                self.style.SUCCESS("Created subscription for {}".format(identity["id"]))
            )
            return sr
        self.stdout.write(
            self.style.NOTICE(
                "{} has active subscriptions, skipping...".format(identity["id"])
            )
        )

    def handle(self, *args, **options):
        contacts = self.extract_contacts_from_workbook(options["file_path"])
        contacts = map(
            lambda c: c._replace(
                msisdn=self.validate_msisdn(c.msisdn),
                language=self.validate_language(c.language),
            ),
            contacts,
        )
        contacts = filter(
            lambda c: c.msisdn is not None and c.language is not None, contacts
        )
        identities = map(
            lambda c: self.create_or_update_identity(c.msisdn, c.language), contacts
        )
        subscriptions = map(lambda i: self.create_subscription_postbirth(i), identities)
        count = sum(1 for s in subscriptions if s is not None)
        self.stdout.write(self.style.SUCCESS("Created {} subscriptions".format(count)))
