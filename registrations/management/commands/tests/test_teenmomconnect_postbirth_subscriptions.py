import io
import random
import string
import tempfile
import uuid
from unittest.mock import call, patch
from urllib.parse import urlencode

import responses
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import TestCase
from openpyxl import Workbook

from registrations.management.commands.teenmomconnect_postbirth_subscriptions import (
    Command,
    Contact,
)
from registrations.models import SubscriptionRequest


def randomword(length):
    """Returns a random string of ascii letters of length `length`"""
    return "".join(random.choice(string.ascii_letters) for _ in range(length))


class TeenMomConnectPostbirthSubscriptionsCommandTests(TestCase):
    @staticmethod
    def generate_workbook(contacts):
        """
        Generates a workbook with the specified `contacts`. Returns the file handler
        of the file where the workbook is stored
        """
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet()
        worksheet.append(["Contact UUID", "Name", "Language", "Phone", "Random"])
        for c in contacts:
            worksheet.append(
                [
                    str(uuid.uuid4()),
                    randomword(10),
                    c.language,
                    c.msisdn,
                    randomword(10),
                ]
            )

        f = tempfile.NamedTemporaryFile(suffix=".xlsx")
        workbook.save(f.name)
        return f

    @staticmethod
    def generate_identity(msisdn, language):
        return {
            "id": "identity-uuid",
            "details": {
                "lang_code": language,
                "default_addr_type": "msisdn",
                "addresses": {"msisdn": {msisdn: {"default": True, "optedout": False}}},
            },
        }

    def test_extract_contacts_from_workbook(self):
        """
        If the workbook is valid, an iterable of Contacts should be returned.
        """
        contacts = [Contact("+27820001001", "eng"), Contact("+27820001002", "xho")]
        f = self.generate_workbook(contacts)
        extracted_contacts = Command.extract_contacts_from_workbook(f.name)
        self.assertEqual(list(extracted_contacts), contacts)

    def test_extract_contacts_from_workbook_multiple_sheets(self):
        """
        If there is more than one sheet in the workbook, then a CommandError should
        be raised
        """
        workbook = Workbook(write_only=True)
        for _ in range(2):
            workbook.create_sheet()
        f = tempfile.NamedTemporaryFile(suffix=".xlsx")
        workbook.save(f.name)
        with self.assertRaises(CommandError) as e:
            list(Command.extract_contacts_from_workbook(f.name))
        self.assertIn("Document can only have a single sheet", str(e.exception))

    def test_extract_contacts_from_workbook_no_phone_header(self):
        """
        If there is no `Phone` header field, then a CommandError should be raised
        """
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet()
        worksheet.append(["Contact UUID", "Name", "Language", "Random"])
        f = tempfile.NamedTemporaryFile(suffix=".xlsx")
        workbook.save(f.name)
        with self.assertRaises(CommandError) as e:
            list(Command.extract_contacts_from_workbook(f.name))
        self.assertIn(
            "Sheet must have a single 'Phone' column header", str(e.exception)
        )

    def test_extract_contacts_from_workbook_no_language_header(self):
        """
        If there is no `Phone` header field, then a CommandError should be raised
        """
        workbook = Workbook(write_only=True)
        worksheet = workbook.create_sheet()
        worksheet.append(["Contact UUID", "Name", "Phone", "Random"])
        f = tempfile.NamedTemporaryFile(suffix=".xlsx")
        workbook.save(f.name)
        with self.assertRaises(CommandError) as e:
            list(Command.extract_contacts_from_workbook(f.name))
        self.assertIn(
            "Sheet must have a single 'Language' column header", str(e.exception)
        )

    def test_extract_contacts_from_workbook_no_header(self):
        """
        If there is no header field, then a CommandError should be raised
        """
        workbook = Workbook(write_only=True)
        workbook.create_sheet()
        f = tempfile.NamedTemporaryFile(suffix=".xlsx")
        workbook.save(f.name)
        with self.assertRaises(CommandError) as e:
            list(Command.extract_contacts_from_workbook(f.name))
        self.assertIn(
            "Sheet must have a single 'Phone' column header", str(e.exception)
        )

    def test_validate_msisdn(self):
        """
        If the msisdn is valid, it should be returned in E164 international format
        """
        command = Command()
        self.assertEqual(command.validate_msisdn("0820001001"), "+27820001001")

    def test_validate_msisdn_not_parsable(self):
        """
        If it's not a parsable msisdn, an error should be logged and `None` should be
        returned
        """
        command = Command()
        command.stdout = io.StringIO()
        self.assertEqual(command.validate_msisdn("gibberish"), None)
        self.assertIn(
            "Invalid phone number gibberish. Skipping...", command.stdout.getvalue()
        )

    def test_validate_msisdn_not_possible(self):
        """
        If it's not a possible msisdn, an error should be logged and `None` should be
        returned
        """
        command = Command()
        command.stdout = io.StringIO()
        self.assertEqual(command.validate_msisdn("+1200012301"), None)
        self.assertIn(
            "Invalid phone number +1200012301. Skipping...", command.stdout.getvalue()
        )

    def test_validate_msisdn_not_valid(self):
        """
        If it's not a valid msisdn, an error should be logged and `None` should be
        returned
        """
        command = Command()
        command.stdout = io.StringIO()
        self.assertEqual(command.validate_msisdn("+12001230101"), None)
        self.assertIn(
            "Invalid phone number +12001230101. Skipping...", command.stdout.getvalue()
        )

    def test_validate_language(self):
        """
        If the language code is valid, it should be returned in the Seed language format
        """
        command = Command()
        self.assertEqual(command.validate_language("eng"), "eng_ZA")

    def test_validate_language_invalid(self):
        """
        If the language code is invalid, an error should be logged and None returned
        """
        command = Command()
        command.stdout = io.StringIO()
        self.assertEqual(command.validate_language("foo"), None)
        self.assertIn(
            "Invalid language code foo. Skipping...", command.stdout.getvalue()
        )

    def test_validate_language_not_in_seed(self):
        """
        If the language code is not in the project's list of languages, an error should
        be logged and None returned
        """
        command = Command()
        command.stdout = io.StringIO()
        self.assertEqual(command.validate_language("arg"), None)
        self.assertIn(
            "Invalid language code arg. Skipping...", command.stdout.getvalue()
        )

    @responses.activate
    def test_create_or_update_identity_exists(self):
        """
        If the identity exists, and the language code is correct, it should be returned
        """
        identity = self.generate_identity("+27820001001", "eng_ZA")
        responses.add(
            responses.GET,
            "http://is/api/v1/identities/search/?{}".format(
                urlencode({"details__addresses__msisdn": "+27820001001"})
            ),
            json={"results": [identity]},
        )
        result = Command.create_or_update_identity("+27820001001", "eng_ZA")
        self.assertEqual(result, identity)
        self.assertEqual(len(responses.calls), 1)

    @responses.activate
    def test_create_or_update_identity_does_not_exist(self):
        """
        If the identity does not exist, it should be created
        """
        identity = self.generate_identity("+27820001001", "eng_ZA")
        responses.add(
            responses.GET,
            "http://is/api/v1/identities/search/?{}".format(
                urlencode({"details__addresses__msisdn": "+27820001001"})
            ),
            json={"results": []},
        )
        responses.add(responses.POST, "http://is/api/v1/identities/", json=identity)
        result = Command.create_or_update_identity("+27820001001", "eng_ZA")
        self.assertEqual(result, identity)
        self.assertEqual(len(responses.calls), 2)

    @responses.activate
    def test_create_or_update_identity_language_incorrect(self):
        """
        If the identity does not have a language, or an incorrect language, it should
        be updated
        """
        identity = self.generate_identity("+27820001001", "eng_ZA")
        identity_incorrect_lang = self.generate_identity("+27820001001", "afr_ZA")
        responses.add(
            responses.GET,
            "http://is/api/v1/identities/search/?{}".format(
                urlencode({"details__addresses__msisdn": "+27820001001"})
            ),
            json={"results": [identity_incorrect_lang]},
        )
        responses.add(
            responses.PATCH, "http://is/api/v1/identities/identity-uuid/", json=identity
        )
        result = Command.create_or_update_identity("+27820001001", "eng_ZA")
        self.assertEqual(result, identity)
        self.assertEqual(len(responses.calls), 2)

    @responses.activate
    def test_create_subscription_postbirth_existing(self):
        """
        If the mother has an existing active subscription, we should log an error and
        not create the new subscription
        """
        command = Command()
        command.stdout = io.StringIO()
        identity = self.generate_identity("+27820001001", "eng_ZA")
        responses.add(
            responses.GET,
            "http://sbm/api/v1/subscriptions/?{}".format(
                urlencode({"identity": "identity-uuid", "active": "True"})
            ),
            json={"results": [{"id": "subscription-uuid"}]},
        )

        command.create_subscription_postbirth(identity)
        self.assertIn(
            "identity-uuid has active subscriptions, skipping...",
            command.stdout.getvalue(),
        )
        self.assertEqual(SubscriptionRequest.objects.count(), 0)

    @responses.activate
    def test_create_subscription_postbirth_non_existing(self):
        """
        If the mother has no existing active subscriptions, we should create a
        subscription
        """
        command = Command()
        command.stdout = io.StringIO()
        identity = self.generate_identity("+27820001001", "eng_ZA")
        responses.add(
            responses.GET,
            "http://sbm/api/v1/subscriptions/?{}".format(
                urlencode({"identity": "identity-uuid", "active": "True"})
            ),
            json={"results": []},
        )
        responses.add(
            responses.GET,
            "http://sbm/api/v1/messageset/?{}".format(
                urlencode({"short_name": "momconnect_postbirth.hw_full.1"})
            ),
            json={"results": [{"id": 1, "default_schedule": 2}]},
        )

        command.create_subscription_postbirth(identity)
        self.assertIn(
            "Created subscription for identity-uuid", command.stdout.getvalue()
        )
        [subreq] = SubscriptionRequest.objects.all()
        self.assertEqual(subreq.identity, "identity-uuid")
        self.assertEqual(subreq.messageset, 1)
        self.assertEqual(subreq.lang, "eng_ZA")
        self.assertEqual(subreq.schedule, 2)

    @patch(
        "registrations.management.commands.teenmomconnect_postbirth_subscriptions."
        "Command.create_or_update_identity"
    )
    @patch(
        "registrations.management.commands.teenmomconnect_postbirth_subscriptions."
        "Command.create_subscription_postbirth"
    )
    def test_run_command(
        self, create_subscription_postbirth, create_or_update_identity
    ):
        """
        Running the command should filter out bad rows, properly format the data of the
        rest, and create identities and subscriptions for them.
        """

        def identity_side_effect(msisdn, lang):
            return {"msisdn": msisdn, "lang": lang}

        create_or_update_identity.side_effect = identity_side_effect
        contacts = [
            Contact("0820001001", "eng"),  # Valid contact
            Contact("invalid", "xho"),  # Invalid phone number
            Contact("0820001001", "arg"),  # Invalid language
            Contact("0820001002", "afr"),  # Valid contact
        ]
        f = self.generate_workbook(contacts)

        out = io.StringIO()
        call_command("teenmomconnect_postbirth_subscriptions", f.name, stdout=out)
        self.assertIn("Invalid phone number invalid. Skipping...", out.getvalue())
        self.assertIn("Invalid language code arg. Skipping...", out.getvalue())
        self.assertIn("Created 2 subscription", out.getvalue())
        create_or_update_identity.assert_has_calls(
            [call("+27820001001", "eng_ZA"), call("+27820001002", "afr_ZA")]
        )
        create_subscription_postbirth.assert_has_calls(
            [
                call({"msisdn": "+27820001001", "lang": "eng_ZA"}),
                call({"msisdn": "+27820001002", "lang": "afr_ZA"}),
            ]
        )
